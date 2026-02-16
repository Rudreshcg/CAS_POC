import os
import json
import http.client
import requests
import re
from flask import Flask, render_template, request, redirect, url_for
import PyPDF2
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date
from flask import send_from_directory
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
class PrefixMiddleware:
    def __init__(self, app):
        self.app = app

    def __call__(self, environ, start_response):
        script_name = environ.get('HTTP_X_SCRIPT_NAME', '')
        if script_name:
            environ['SCRIPT_NAME'] = script_name
            path_info = environ['PATH_INFO']
            if path_info.startswith(script_name):
                environ['PATH_INFO'] = path_info[len(script_name):]
        return self.app(environ, start_response)

app = Flask(__name__)
app.wsgi_app = PrefixMiddleware(app.wsgi_app)

RECAPTCHA_SECRET_KEY = '6LdIebwqAAAAAKHDuP78b-y_cXSryIxgR_1_D8B0'
RECAPTCHA_SITE_KEY = '6LdIebwqAAAAAB-BXd1-YlfIZ8193I5YBihLfTeq'

# Set the upload folder (make sure this directory exists)
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def send_email(subject, body):
    # Create the multipart container
    msg = MIMEMultipart()
    msg['From'] = "kadidalscmmax@gmail.com"
    msg['To'] = "akshay.kadidal@scmmax.com, rama@scmmax.com, sachin.mathur@scmmax.com"
    msg['Subject'] = subject

    # Attach the body of the email to the MIMEText container
    msg.attach(MIMEText(body, 'plain'))

    try:
        # Set up the SMTP server and connect securely
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()  # Secure the connection
        server.login("kadidalscmmax@gmail.com", "zqww bkfp eran ffsl")
        text = msg.as_string()
        server.sendmail("kadidalscmmax@gmail.com", "akshay.kadidal@scmmax.com", text)
        server.quit()
        print("Email sent successfully!")
    except Exception as e:
        print(f"Failed to send email: {e}")

def is_business_email(email):
    """
    Validate if the email is a business email (not from free email providers)
    Returns (is_valid, message)
    """
    domain = email.lower().split('@')[-1]
    
    FREE_EMAIL_DOMAINS = {
    'gmail.com',
    'yahoo.com',
    'hotmail.com',
    'outlook.com',
    'aol.com',
    'icloud.com',
    'protonmail.com',
    'zoho.com',
    'mail.com',
    'yandex.com',
    'gmx.com',
    'live.com',
    'msn.com',
    'yahoo.co.uk',
    'yahoo.co.in',
    'yahoo.co.jp',
    'me.com',
    'mac.com'
}
    
    # Check if domain is in the blocked list
    if domain in FREE_EMAIL_DOMAINS:
        return False, "Please use your business email address."
    
    # Additional check for subdomains of free email providers
    for free_domain in FREE_EMAIL_DOMAINS:
        if domain.endswith(f'.{free_domain}'):
            return False, "Please use your business email address."
            
    return True, "Valid business email"


def g_search(query):
    """
    Perform a Google API search using the specified query.
    """
    conn = http.client.HTTPSConnection("google.serper.dev")
    X_API_KEY = '362de99dbd1b66f05282fd62aa1aa832001047a2'  # Ensure environment variable is set

    payload = json.dumps({
        "q": query
    })

    headers = {
        'X-API-KEY': X_API_KEY,
        'Content-Type': 'application/json'
    }

    conn.request("POST", "/search", payload, headers)
    res = conn.getresponse()
    data = res.read()
    search_results = data.decode("utf-8")
    return search_results


def search_material_code(raw_material):
    """
    Search for the HS code of the given raw material.
    """
    search_result = g_search(f"What is the HS code for {raw_material}?")
    return search_result


def search_material_price(raw_material):
    """
    Search for the current market price of the given raw material in USD.
    """
    search_result = g_search(f"What is the current market price for {raw_material} in USD?")
    return search_result


def process_HS_Code(input_text):
    """
    Extract details from input text and search for additional information like HS code and price.
    """
    url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash-latest:generateContent"
    api_key = 'AIzaSyAkOhIn905jdz6t0qcK2xR1UOp5z-bKeUM'  # Replace with your actual API key

    headers = {
        "Content-Type": "application/json"
    }

    data = {
        "contents": [
            {
                "parts": [
                    {
                         "text": f"You are an agent helping industrial chemical procurement. Read the JSON and return the appropriate HS code as an integer {input_text}\nReturn only one code and return it in integer. Do not return more than one value."
                    }
                ]
            }
        ]
    }

    # Add the API key to the URL
    url_with_key = f"{url}?key={api_key}"

    # Send the POST request
    response = requests.post(url_with_key, headers=headers, json=data)

    if response.status_code != 200:
        return {"error": f"API error: {response.status_code}"}

    try:
        content = response.json()
        parts = content.get('candidates', [])[0].get('content', {}).get('parts', [])
        json_text = parts[0]['text'] if parts else ""

        return(json_text)

    except requests.exceptions.RequestException as e:
        return {"error": f"API request error: {e}"}
    except json.JSONDecodeError as e:
        return {"error": f"JSON decoding error: {e}"}
    except Exception as e:
        return {"error": f"An unexpected error occurred: {e}"}

def process_price(input_text):
    """
    Extract details from input text and search for additional information like HS code and price.
    """
    url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash-latest:generateContent"
    api_key = 'AIzaSyAkOhIn905jdz6t0qcK2xR1UOp5z-bKeUM'  # Replace with your actual API key

    headers = {
        "Content-Type": "application/json"
    }

    data = {
        "contents": [
            {
                "parts": [
                    {
                         "text": f" You are an agent helping industrial chemical procurement. read the json and return the appropriate unit price as integer {input_text}\n Return only one price and return it as an integer. Do not return more than one value"
                    }
                ]
            }
        ]
    }

    # Add the API key to the URL
    url_with_key = f"{url}?key={api_key}"

    # Send the POST request
    response = requests.post(url_with_key, headers=headers, json=data)

    if response.status_code != 200:
        return {"error": f"API error: {response.status_code}"}

    try:
        content = response.json()
        parts = content.get('candidates', [])[0].get('content', {}).get('parts', [])
        json_text = parts[0]['text'] if parts else ""
        
        return(json_text)


    except requests.exceptions.RequestException as e:
        return {"error": f"API request error: {e}"}
    except json.JSONDecodeError as e:
        return {"error": f"JSON decoding error: {e}"}
    except Exception as e:
        return {"error": f"An unexpected error occurred: {e}"}

def process_text(input_text):
    """
    Extract details from input text and search for additional information like HS code and price.
    """
    url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash-latest:generateContent"
    api_key = 'AIzaSyAkOhIn905jdz6t0qcK2xR1UOp5z-bKeUM'  # Replace with your actual API key

    headers = {
        "Content-Type": "application/json"
    }

    data = {
        "contents": [
            {
                "parts": [
                    {
                        "text": f"Extract the details of Raw Material, Vendor Name, minimum Price per Unit, Units, and Currency from following text. \n text: {input_text}\nGenerate a JSON object. Include the following keys: \"Raw_Material\" (string), \"Vendor_Name\" (string), \"Price_per_Unit\" (integer), \"Currency\" (string), and \"Country_of_Origin\" (an array of strings)."
                    }
                ]
            }
        ]
    }

    # Add the API key to the URL
    url_with_key = f"{url}?key={api_key}"

    # Send the POST request
    response = requests.post(url_with_key, headers=headers, json=data)

    if response.status_code != 200:
        return {"error": f"API error: {response.status_code}"}

    try:
        content = response.json()
        parts = content.get('candidates', [])[0].get('content', {}).get('parts', [])
        json_text = parts[0]['text'] if parts else ""

        json_pattern = r"```json\n(.*?)\n```"
        match = re.search(json_pattern, json_text, re.DOTALL)

        if match:
            json_data = match.group(1)
            details = json.loads(json_data)

            # Search for additional details
            raw_material = details.get("Raw_Material", "")
            hs_code_details = search_material_code(raw_material)
            price_details = search_material_price(raw_material)

            price = process_price(price_details)
            hs_code = process_HS_Code(hs_code_details)

            return {
                "country_of_origin": details.get("Country_of_Origin", ""),
                "pricing": details.get("Price_per_Unit", []),
                "raw_material": raw_material,
                "vendor_name": details.get("Vendor_Name", ""),
                "currency": details.get("Currency", ""),
                "hs_code": hs_code,
                "price": price
            }
        else:
            return {"error": "No JSON data found in the API response"}

    except requests.exceptions.RequestException as e:
        return {"error": f"API request error: {e}"}
    except json.JSONDecodeError as e:
        return {"error": f"JSON decoding error: {e}"}
    except Exception as e:
        return {"error": f"An unexpected error occurred: {e}"}

def is_float(element):
    try:
        float(element)
        return True
    except ValueError:
        return False
    
def analyze_quotes(processed_files):
    """
    Analyze multiple quotes and provide recommendations
    """
    if not processed_files:
        return None
    
    # Find the lowest price
    #prices = [float(quote.get('pricing', float('inf'))) for quote in processed_files]
    prices = []
    for quote in processed_files:
        pricing = quote.get('pricing')
        # Check if pricing is None
        if pricing is None:
            prices.append(float('inf'))
        elif isinstance(pricing, (int, float)):
            # Directly convert single numeric values to float
            prices.append(float(pricing))
        elif isinstance(pricing, str):
            # Attempt to convert string to float
            try:
                prices.append(float(pricing))
            except ValueError:
                # If conversion fails, use float('inf')
                prices.append(float('inf'))
        elif isinstance(pricing, list):
            # If pricing is a list, attempt to find the minimum valid float value
            try:
                # Filter the list to only include valid floats, ignoring non-numeric entries
                valid_numbers = [float(num) for num in pricing if isinstance(num, (int, float, str)) and is_float(num)]
                if valid_numbers:
                    # If there are valid numbers, append the minimum
                    prices.append(min(valid_numbers))
                else:
                    # If no valid numbers, use float('inf')
                    prices.append(float('inf'))
            except ValueError:
                # Handle any conversion errors
                prices.append(float('inf'))
        else:
            # Default to inf if pricing is of any other unsupported type
            prices.append(float('inf'))

    
    min_price = min(prices)
    best_price_index = prices.index(min_price)
    
    # Compare prices and create analysis
    analysis = {
        'best_choice': processed_files[best_price_index],
        'price_comparison': [],
        'recommendation': ''
    }
    
    # Calculate price differences
    for i, quote in enumerate(processed_files):
        pricing = quote.get('pricing', 0)
        if isinstance(pricing, list):
            prices = [float(p) for p in pricing if is_float(p)]
            price = min(prices) if prices else 0
        elif pricing is None:
            price = 0
        else:
            price = float(quote.get('pricing', 0))
            difference = ((price - min_price) / min_price) * 100 if min_price > 0 else 0
            analysis['price_comparison'].append({
                'vendor': quote.get('vendor_name'),
                'price': price,
                'difference_percentage': round(difference, 2)
            })
    
    # Generate recommendation
    best_vendor = processed_files[best_price_index].get('vendor_name')
    best_price = processed_files[best_price_index].get('price')
    raw_material = processed_files[best_price_index].get('raw_material')
    
    analysis['recommendation'] = f"Based on the price comparison, {best_vendor} offers the best value for {raw_material} at ${best_price}. "
    
    # Add additional insights
    if len(processed_files) > 1:
        second_best = sorted(analysis['price_comparison'], key=lambda x: x['price'])[1]
        price_difference = second_best['difference_percentage']
        analysis['recommendation'] += f"This is {price_difference}% lower than the next best option from {second_best['vendor']}."
    
    return analysis

def convert_currency(amount, from_currency, to_currency, api_key='1753eae9e2239f2071421196'):
    """
    Convert amount from one currency to another using exchangerate-api.com.

    :param amount: float, the amount to convert
    :param from_currency: str, the currency code to convert from (e.g., 'USD')
    :param to_currency: str, the currency code to convert to (e.g., 'EUR')
    :param api_key: str, your API key for exchangerate-api.com
    :return: float, the converted amount
    """
    url = f"https://v6.exchangerate-api.com/v6/{api_key}/pair/{from_currency}/{to_currency}/{amount}"
    
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        if data['result'] == 'success':
            return data['conversion_result']
        else:
            raise Exception(f"Error in fetching currency data: {data['error-type']}")
    else:
        response.raise_for_status()


def create_excel(processed_data):
    wb = load_workbook('Comparison.xlsx')
    ws = wb.active

    data = processed_data[0] 
    # update standard values
    ws['D4'] = data.get('raw_material', '')
    
    # update date
    today = date.today()
    d1 = today.strftime('%d/%m/%Y')
    ws['G4'] = "Date Prepared: " + d1

    # Update other values
    ws['D5'] = data.get('hs_code', '')
    ws['D14'] = data.get('price', '')
    ws['B14'] = 'Market Price'



    # Add data rows
    row_index = 23
    column_index=4

    for data in processed_data:
            ws.cell(row=row_index, column=column_index, value= data.get('vendor_name', ''))
            ws.cell(row=row_index+1, column=column_index, value= data.get('pricing', ''))
            USD_Price = data.get('pricing', '')
            INR_Price = convert_currency(USD_Price, 'USD','INR')
            ws.cell(row=row_index+2, column=column_index, value= INR_Price)
            Gst_rate = ws.cell(row=row_index+3, column=3).value
            igst_rate= ws.cell(row=row_index+4, column=3).value
            ws.cell(row=row_index+3, column=column_index, value= INR_Price*Gst_rate)
            ws.cell(row=row_index+4, column=column_index, value= INR_Price*igst_rate)
            ws.cell(row=row_index+15, column=column_index, value= ", ".join(data.get('country_of_origin', [])))
            column_index = column_index + 1


    filename = 'quote_comparison.xlsx'
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    wb.save(filepath)
    return filename

@app.route('/')
def home():
    """Render the home landing page"""
    return render_template('index.html', active_tab='home')

@app.route('/login')
def login():
    """Render the login page"""
    return render_template('login.html')

@app.route('/demo')
def demo():
    """Render the demo page"""
    return render_template('demo.html')

@app.route('/login', methods=['POST'])
def login_post():
    """Handle login form submission TBD"""
    # Add your authentication logic here
    # For now, just redirect to demo page
    return redirect(url_for('demo'))

@app.route('/apollo')
def apollo():
    """Render the apollo page"""
    return render_template('index.html', active_tab='apollo')

@app.route('/settings')
def settings():
    """Render the settings page"""
    return render_template('index.html', active_tab='settings')

@app.route('/contact')
def contact():
    """Render the contact page"""
    return render_template('index.html', active_tab='contact')

@app.route('/process-quotes', methods=['POST'])
def process_quotes():
    """Handle quote processing"""
    processed_data = []
    analysis = None

    if 'files[]' not in request.files:
        return render_template('index.html', error="No files selected", active_tab='apollo')

    files = request.files.getlist('files[]')
    
    if not files or files[0].filename == '':
        return render_template('index.html', error="No selected files", active_tab='apollo')

    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

    for file in files:
        if file and file.filename.lower().endswith('.pdf'):
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            try:
                file.save(filepath)
                with open(filepath, 'rb') as pdf_file:
                    pdf_reader = PyPDF2.PdfReader(pdf_file)
                    full_text = "".join([page.extract_text() for page in pdf_reader.pages])
                
                result = process_text(full_text)
                if not isinstance(result, dict) or 'error' not in result:
                    processed_data.append(result)
                    
            except Exception as e:
                return render_template('index.html', error=f"Error processing {file.filename}: {str(e)}", active_tab='apollo')
            finally:
                if os.path.exists(filepath):
                    os.remove(filepath)
        else:
            return render_template('index.html', error="Please upload PDF files only", active_tab='apollo')

    if processed_data:
        analysis = analyze_quotes(processed_data)
        excel_filename = create_excel(processed_data)
        return render_template('index.html', 
                             processed_data=processed_data, 
                             analysis=analysis, 
                             excel_filename=excel_filename,
                             active_tab='apollo')

    return render_template('index.html', active_tab='apollo')

@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)


@app.route('/request-demo', methods=['POST'])
def request_demo():
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        company = request.form.get('company')
        recaptcha_response = request.form.get('g-recaptcha-response')
        
        # Validate email
        is_valid, message = is_business_email(email)
        
        if not is_valid:
            return render_template('index.html', 
                                active_tab='home',
                                error=message)
        

        # --- RECAPTCHA VERIFICATION ---
        if not recaptcha_response:
            return render_template('index.html', 
                                active_tab='home', 
                                error="Please complete the reCAPTCHA.")

        # Verify reCAPTCHA with Google's servers
        recaptcha_url = "https://www.google.com/recaptcha/api/siteverify"
        recaptcha_data = {
            'secret': '6LdIebwqAAAAAKHDuP78b-y_cXSryIxgR_1_D8B0',  
            'response': recaptcha_response
        }

        try:
          response = requests.post(recaptcha_url, data=recaptcha_data)
          response.raise_for_status() # Raise an HTTPError for bad responses
          recaptcha_result = response.json()
        except requests.exceptions.RequestException as e:
            return render_template('index.html',
                           active_tab='home',
                           error=f"Error communicating with reCAPTCHA service: {str(e)}")

        if not recaptcha_result.get('success'):
           return render_template('index.html', 
                                active_tab='home',
                                error="reCAPTCHA verification failed. Please try again.")


        # Create email subject and body
        subject = 'New Demo Request from SCMmax'
        body = f"""
        New demo request received:
        
        Name: {name}
        Email: {email}
        Company: {company}
        """
        
        try:
            send_email(subject, body)
            return render_template('index.html', 
                                active_tab='home',
                                success="Thank you for your request. We'll be in touch soon!")
        except Exception as e:
            return render_template('index.html', 
                                active_tab='home',
                                error=f"An error occurred: {str(e)}")

    return render_template('index.html', active_tab='home')

@app.route('/send-message', methods=['POST'])
def send_message():
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        message = request.form.get('message')
        
        # Validate email
        is_valid, email_message = is_business_email(email)
        
        if not is_valid:
            return render_template('index.html', 
                                active_tab='contact',
                                error=email_message)
        
        # Create email subject and body
        subject = 'New Contact Form Message from SCMmax'
        body = f"""
        New message received:
        
        Name: {name}
        Email: {email}
        Message:
        {message}
        """
        
        try:
            send_email(subject, body)
            return render_template('index.html', 
                                active_tab='contact',
                                success="Thank you for your message. We'll get back to you soon!")
        except Exception as e:
            return render_template('index.html', 
                                active_tab='contact',
                                error=f"An error occurred: {str(e)}")

    return render_template('index.html', active_tab='contact')

@app.route('/debug')
def debug():
    import os
    return {
        'headers': dict(request.headers),
        'url_for_static': url_for('static', filename='img/Astro3.png'),
        'script_root': request.script_root,
        'url_root': request.url_root,
        'environ_script_name': request.environ.get('SCRIPT_NAME', 'NOT_SET'),
        'environ_path_info': request.environ.get('PATH_INFO', 'NOT_SET'),
        'environ_http_x_script_name': request.environ.get('HTTP_X_SCRIPT_NAME', 'NOT_SET')
    }

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5003)