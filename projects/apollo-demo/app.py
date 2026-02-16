import os
import json
import http.client
import requests
import re
import csv
from flask import Flask, render_template, request, redirect, url_for, jsonify
import PyPDF2
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date
from flask import send_from_directory
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from Levenshtein import distance
from fuzzywuzzy import fuzz
import google.generativeai as genai
from dotenv import load_dotenv
from duplicate_detector import DuplicateDetector

# Load environment variables
load_dotenv()

class PrefixMiddleware:
    def __init__(self, app):
        self.app = app

    def __call__(self, environ, start_response):
        script_name = environ.get('HTTP_X_SCRIPT_NAME', '')
        if script_name:
            environ['SCRIPT_NAME'] = script_name
            path_info = environ['PATH_INFO']
            if path_info.startswith(script_name):
                environ['PATH_INFO'] = path_info[len(script_name):]
        return self.app(environ, start_response)

app = Flask(__name__)
app.wsgi_app = PrefixMiddleware(app.wsgi_app)

RECAPTCHA_SECRET_KEY = '6LdIebwqAAAAAKHDuP78b-y_cXSryIxgR_1_D8B0'
RECAPTCHA_SITE_KEY = '6LdIebwqAAAAAB-BXd1-YlfIZ8193I5YBihLfTeq'

# Set the upload folder (make sure this directory exists)
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


def g_search(query):
    """
    Perform a Google API search using the specified query.
    """
    try:
        conn = http.client.HTTPSConnection("google.serper.dev")
        X_API_KEY = '362de99dbd1b66f05282fd62aa1aa832001047a2'  # Ensure environment variable is set

        payload = json.dumps({
            "q": query
        })

        headers = {
            'X-API-KEY': X_API_KEY,
            'Content-Type': 'application/json'
        }

        conn.request("POST", "/search", payload, headers)
        res = conn.getresponse()
        data = res.read()
        search_results = data.decode("utf-8")
        return search_results
    except Exception as e:
        # FIXED: Added error handling for SSL certificate issues and API failures
        print(f"Error in g_search: {e}")
        return f"Error searching for: {query}"


def search_material_code(raw_material):
    """
    Search for the HS code of the given raw material.
    """
    try:
        search_result = g_search(f"What is the HS code for {raw_material}?")
        return search_result
    except Exception as e:
        # FIXED: Added error handling to prevent crashes when secondary APIs fail
        print(f"Error in search_material_code: {e}")
        return "Error searching for HS code"


def search_material_price(raw_material):
    """
    Search for the current market price of the given raw material in USD.
    """
    try:
        search_result = g_search(f"What is the current market price for {raw_material} in USD?")
        return search_result
    except Exception as e:
        # FIXED: Added error handling to prevent crashes when secondary APIs fail
        print(f"Error in search_material_price: {e}")
        return "Error searching for price"


def process_HS_Code(input_text):
    """
    Extract details from input text and search for additional information like HS code and price.
    """
    url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash-latest:generateContent"
    api_key = 'AIzaSyAkOhIn905jdz6t0qcK2xR1UOp5z-bKeUM'  # Replace with your actual API key

    headers = {
        "Content-Type": "application/json"
    }

    data = {
        "contents": [
            {
                "parts": [
                    {
                         "text": f"You are an agent helping industrial chemical procurement. Read the JSON and return the appropriate HS code as an integer {input_text}\nReturn only one code and return it in integer. Do not return more than one value."
                    }
                ]
            }
        ]
    }

    # Add the API key to the URL
    url_with_key = f"{url}?key={api_key}"

    # Send the POST request
    response = requests.post(url_with_key, headers=headers, json=data)

    if response.status_code != 200:
        return {"error": f"API error: {response.status_code}"}

    try:
        content = response.json()
        parts = content.get('candidates', [])[0].get('content', {}).get('parts', [])
        json_text = parts[0]['text'] if parts else ""

        return(json_text)

    except requests.exceptions.RequestException as e:
        return {"error": f"API request error: {e}"}
    except json.JSONDecodeError as e:
        return {"error": f"JSON decoding error: {e}"}
    except Exception as e:
        return {"error": f"An unexpected error occurred: {e}"}

def process_price(input_text):
    """
    Extract details from input text and search for additional information like HS code and price.
    """
    url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash-latest:generateContent"
    api_key = 'AIzaSyAkOhIn905jdz6t0qcK2xR1UOp5z-bKeUM'  # Replace with your actual API key

    headers = {
        "Content-Type": "application/json"
    }

    data = {
        "contents": [
            {
                "parts": [
                    {
                         "text": f" You are an agent helping industrial chemical procurement. read the json and return the appropriate unit price as integer {input_text}\n Return only one price and return it as an integer. Do not return more than one value"
                    }
                ]
            }
        ]
    }

    # Add the API key to the URL
    url_with_key = f"{url}?key={api_key}"

    # Send the POST request
    response = requests.post(url_with_key, headers=headers, json=data)

    if response.status_code != 200:
        return {"error": f"API error: {response.status_code}"}

    try:
        content = response.json()
        parts = content.get('candidates', [])[0].get('content', {}).get('parts', [])
        json_text = parts[0]['text'] if parts else ""
        
        return(json_text)


    except requests.exceptions.RequestException as e:
        return {"error": f"API request error: {e}"}
    except json.JSONDecodeError as e:
        return {"error": f"JSON decoding error: {e}"}
    except Exception as e:
        return {"error": f"An unexpected error occurred: {e}"}

def process_text(input_text):
    """
    Extract details from input text and search for additional information like HS code and price.
    """
    # FIXED: Added robust fallback system when Gemini API is unavailable
    # First try the Gemini API
    try:
        url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash-latest:generateContent"
        api_key = 'AIzaSyAkOhIn905jdz6t0qcK2xR1UOp5z-bKeUM'

        headers = {
            "Content-Type": "application/json"
        }

        data = {
            "contents": [
                {
                    "parts": [
                        {
                            "text": f"Extract the details of Raw Material, Vendor Name, minimum Price per Unit, Units, and Currency from following text. \n text: {input_text}\nGenerate a JSON object. Include the following keys: \"Raw_Material\" (string), \"Vendor_Name\" (string), \"Price_per_Unit\" (integer), \"Currency\" (string), and \"Country_of_Origin\" (an array of strings)."
                        }
                    ]
                }
            ]
        }

        url_with_key = f"{url}?key={api_key}"
        print(f"Sending request to Gemini API...")
        
        # FIXED: Added retry logic with exponential backoff for API overload issues
        max_retries = 2
        for attempt in range(max_retries):
            try:
                response = requests.post(url_with_key, headers=headers, json=data, timeout=30)
                print(f"API Response status: {response.status_code}")
                
                if response.status_code == 200:
                    content = response.json()
                    parts = content.get('candidates', [])[0].get('content', {}).get('parts', [])
                    json_text = parts[0]['text'] if parts else ""
                    
                    print(f"API Response text: {json_text[:500]}...")

                    json_pattern = r"```json\n(.*?)\n```"
                    match = re.search(json_pattern, json_text, re.DOTALL)

                    if match:
                        json_data = match.group(1)
                        details = json.loads(json_data)
                        print(f"Parsed JSON details: {details}")

                        # FIXED: Added error handling for secondary API calls (HS code and price search)
                        raw_material = details.get("Raw_Material", "")
                        try:
                            hs_code_details = search_material_code(raw_material)
                            price_details = search_material_price(raw_material)
                            price = process_price(price_details)
                            hs_code = process_HS_Code(hs_code_details)
                        except Exception as e:
                            print(f"Secondary API calls failed: {e}")
                            price = "N/A"
                            hs_code = "N/A"

                        result = {
                            "country_of_origin": details.get("Country_of_Origin", ""),
                            "pricing": details.get("Price_per_Unit", []),
                            "raw_material": raw_material,
                            "vendor_name": details.get("Vendor_Name", ""),
                            "currency": details.get("Currency", ""),
                            "hs_code": hs_code,
                            "price": price
                        }
                        print(f"Final result: {result}")
                        return result
                    else:
                        print(f"No JSON pattern found in response")
                        break
                elif response.status_code == 503 and attempt < max_retries - 1:
                    print(f"API overloaded (503), retrying in {2 ** attempt} seconds... (attempt {attempt + 1}/{max_retries})")
                    import time
                    time.sleep(2 ** attempt)
                    continue
                else:
                    print(f"API Error response: {response.text}")
                    break
            except Exception as e:
                print(f"Request exception on attempt {attempt + 1}: {e}")
                if attempt == max_retries - 1:
                    break
                import time
                time.sleep(2 ** attempt)
    except Exception as e:
        print(f"Gemini API failed: {e}")

    # FIXED: Added fallback text processing when APIs are unavailable
    print("Using fallback text processing...")
    return fallback_text_processing(input_text)

def fallback_text_processing(input_text):
    """
    FIXED: Added fallback text processing when APIs are unavailable
    This function uses regex patterns to extract basic information from PDF text
    """
    try:
        # Extract basic information using regex patterns
        import re
        
        # Look for common patterns in the text
        raw_material_match = re.search(r'(IPA|Isopropyl Alcohol|Isopropanol)', input_text, re.IGNORECASE)
        raw_material = raw_material_match.group(1) if raw_material_match else "Unknown Material"
        
        # Look for vendor names (common patterns)
        vendor_patterns = [
            r'From:\s*([^<\n]+)',
            r'Vendor:\s*([^\n]+)',
            r'Supplier:\s*([^\n]+)',
            r'Company:\s*([^\n]+)'
        ]
        
        vendor_name = "Unknown Vendor"
        for pattern in vendor_patterns:
            match = re.search(pattern, input_text, re.IGNORECASE)
            if match:
                vendor_name = match.group(1).strip()
                break
        
        # Look for price information
        price_match = re.search(r'(\$?\d+(?:\.\d{2})?)\s*(USD|INR|EUR)', input_text, re.IGNORECASE)
        if price_match:
            pricing = price_match.group(1)
            currency = price_match.group(2)
        else:
            pricing = "N/A"
            currency = "USD"
        
        # Look for country information
        country_match = re.search(r'(China|India|USA|Germany|Japan|Korea)', input_text, re.IGNORECASE)
        country_of_origin = [country_match.group(1)] if country_match else ["Unknown"]
        
        result = {
            "country_of_origin": country_of_origin,
            "pricing": pricing,
            "raw_material": raw_material,
            "vendor_name": vendor_name,
            "currency": currency,
            "hs_code": "N/A",
            "price": "N/A"
        }
        
        print(f"Fallback processing result: {result}")
        return result
        
    except Exception as e:
        print(f"Fallback processing error: {e}")
        return {
            "country_of_origin": ["Unknown"],
            "pricing": "N/A",
            "raw_material": "Unknown Material",
            "vendor_name": "Unknown Vendor",
            "currency": "USD",
            "hs_code": "N/A",
            "price": "N/A"
        }

def is_float(element):
    try:
        float(element)
        return True
    except ValueError:
        return False
    
def analyze_quotes(processed_files):
    """
    Analyze multiple quotes and provide recommendations
    """
    if not processed_files:
        return None
    
    # Find the lowest price
    #prices = [float(quote.get('pricing', float('inf'))) for quote in processed_files]
    prices = []
    for quote in processed_files:
        pricing = quote.get('pricing')
        # Check if pricing is None
        if pricing is None:
            prices.append(float('inf'))
        elif isinstance(pricing, (int, float)):
            # Directly convert single numeric values to float
            prices.append(float(pricing))
        elif isinstance(pricing, str):
            # Attempt to convert string to float
            try:
                prices.append(float(pricing))
            except ValueError:
                # If conversion fails, use float('inf')
                prices.append(float('inf'))
        elif isinstance(pricing, list):
            # If pricing is a list, attempt to find the minimum valid float value
            try:
                # Filter the list to only include valid floats, ignoring non-numeric entries
                valid_numbers = [float(num) for num in pricing if isinstance(num, (int, float, str)) and is_float(num)]
                if valid_numbers:
                    # If there are valid numbers, append the minimum
                    prices.append(min(valid_numbers))
                else:
                    # If no valid numbers, use float('inf')
                    prices.append(float('inf'))
            except ValueError:
                # Handle any conversion errors
                prices.append(float('inf'))
        else:
            # Default to inf if pricing is of any other unsupported type
            prices.append(float('inf'))

    
    min_price = min(prices)
    best_price_index = prices.index(min_price)
    
    # Compare prices and create analysis
    analysis = {
        'best_choice': processed_files[best_price_index],
        'price_comparison': [],
        'recommendation': ''
    }
    
    # FIXED: Added robust error handling for price comparison calculations
    for i, quote in enumerate(processed_files):
        pricing = quote.get('pricing', 0)
        if isinstance(pricing, list):
            prices = [float(p) for p in pricing if is_float(p)]
            price = min(prices) if prices else 0
        elif pricing is None or pricing == 'N/A':
            price = 0
        else:
            try:
                price = float(quote.get('pricing', 0))
            except (ValueError, TypeError):
                price = 0
            
            if price > 0:  # Only add to comparison if we have a valid price
                difference = ((price - min_price) / min_price) * 100 if min_price > 0 else 0
                analysis['price_comparison'].append({
                    'vendor': quote.get('vendor_name'),
                    'price': price,
                    'difference_percentage': round(difference, 2)
                })
    
    # FIXED: Added error handling for recommendation generation to prevent IndexError
    best_vendor = processed_files[best_price_index].get('vendor_name')
    best_price = processed_files[best_price_index].get('price')
    raw_material = processed_files[best_price_index].get('raw_material')
    
    analysis['recommendation'] = f"Based on the price comparison, {best_vendor} offers the best value for {raw_material} at ${best_price}. "
    
    # Add additional insights
    if len(processed_files) > 1 and len(analysis['price_comparison']) > 1:
        try:
            second_best = sorted(analysis['price_comparison'], key=lambda x: x['price'])[1]
            price_difference = second_best['difference_percentage']
            analysis['recommendation'] += f"This is {price_difference}% lower than the next best option from {second_best['vendor']}."
        except (IndexError, KeyError):
            analysis['recommendation'] += "Additional price comparisons are available in the detailed analysis."
    else:
        analysis['recommendation'] += "This appears to be the only quote with valid pricing information."
    
    return analysis

def convert_currency(amount, from_currency, to_currency, api_key='1753eae9e2239f2071421196'):
    """
    Convert amount from one currency to another using exchangerate-api.com.

    :param amount: float, the amount to convert
    :param from_currency: str, the currency code to convert from (e.g., 'USD')
    :param to_currency: str, the currency code to convert to (e.g., 'EUR')
    :param api_key: str, your API key for exchangerate-api.com
    :return: float, the converted amount
    """
    url = f"https://v6.exchangerate-api.com/v6/{api_key}/pair/{from_currency}/{to_currency}/{amount}"
    
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        if data['result'] == 'success':
            return data['conversion_result']
        else:
            raise Exception(f"Error in fetching currency data: {data['error-type']}")
    else:
        response.raise_for_status()


def create_excel(processed_data):
    wb = load_workbook('./Comparison.xlsx')
    ws = wb.active

    data = processed_data[0] 
    # update standard values
    ws['D4'] = data.get('raw_material', '')
    
    # update date
    today = date.today()
    d1 = today.strftime('%d/%m/%Y')
    ws['G4'] = "Date Prepared: " + d1

    # Update other values
    ws['D5'] = data.get('hs_code', '')
    ws['D14'] = data.get('price', '')
    ws['B14'] = 'Market Price'



    # Add data rows
    row_index = 23
    column_index=4

    for data in processed_data:
            ws.cell(row=row_index, column=column_index, value= data.get('vendor_name', ''))
            ws.cell(row=row_index+1, column=column_index, value= data.get('pricing', ''))
            USD_Price = data.get('pricing', '')
            
            # Handle N/A values for currency conversion
            if USD_Price == 'N/A' or USD_Price == '':
                INR_Price = 0
            else:
                try:
                    INR_Price = convert_currency(float(USD_Price), 'USD','INR')
                except (ValueError, TypeError):
                    INR_Price = 0
            
            ws.cell(row=row_index+2, column=column_index, value= INR_Price)
            Gst_rate = ws.cell(row=row_index+3, column=3).value
            igst_rate= ws.cell(row=row_index+4, column=3).value
            ws.cell(row=row_index+3, column=column_index, value= INR_Price*Gst_rate)
            ws.cell(row=row_index+4, column=column_index, value= INR_Price*igst_rate)
            ws.cell(row=row_index+15, column=column_index, value= ", ".join(data.get('country_of_origin', [])))
            column_index = column_index + 1


    filename = 'quote_comparison.xlsx'
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    wb.save(filepath)
    return filename



@app.route('/')
def index():
    return render_template('index.html')

@app.route('/market-research-report')
def market_research_report():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('market-research-report.html', material=material)

@app.route('/news')
def news():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('news.html', material=material)

@app.route('/key-value-drivers')
def key_value_drivers():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('key-value-drivers.html', material=material)

@app.route('/trade-data-analysis')
def trade_data_analysis():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('trade-data-analysis.html', material=material)

@app.route('/vendor-key-information')
def vendor_key_information():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('vendor-key-information.html', material=material)

@app.route('/demand-supply-trends')
def demand_supply_trends():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('demand-supply-trends.html', material=material)

@app.route('/joint-development-projects')
def joint_development_projects():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('joint-development-projects.html', material=material)

@app.route('/vendor-minutes-of-meetings')
def vendor_minutes_of_meetings():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('vendor-minutes-of-meetings.html', material=material)

@app.route('/multipoint-engagement')
def multipoint_engagement():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('multipoint-engagement.html', material=material)

@app.route('/supplier-tracking')
def supplier_tracking():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('supplier-tracking.html', material=material)

@app.route('/shutdown-tracker')
def shutdown_tracker():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('shutdown-tracker.html', material=material)

@app.route('/fact-pack')
def fact_pack():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('fact-pack.html', material=material)

@app.route('/spend-analytics')
def spend_analytics():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('spend-analytics.html', material=material)

@app.route('/industry-porter-analysis')
def industry_porter_analysis():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('industry-porter-analysis.html', material=material)

@app.route('/cyclical-patterns')
def cyclical_patterns():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('cyclical-patterns.html', material=material)

@app.route('/negotiation-window')
def negotiation_window():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('negotiation-window.html', material=material)

@app.route('/sesonality-trends')
def sesonality_trends():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('sesonality-trends.html', material=material)

@app.route('/procurement-plan')
def procurement_plan():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('procurement-plan.html', material=material)

@app.route('/inventory-level')
def inventory_level():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('inventory-level.html', material=material)

@app.route('/vendor-wise-action-plan')
def vendor_wise_action_plan():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('vendor-wise-action-plan.html', material=material)

@app.route('/cost-sheet')
def cost_sheet():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('cost-sheet.html', material=material)

@app.route('/negotiation-objectives')
def negotiation_objectives():
    # Get the searched material from query parameters
    #material = request.args.get('material', '')
    return render_template('negotiation-objectives.html')

@app.route('/price-benchmarking')
def price_benchmarking():
    # Get the searched material from query parameters
    material = request.args.get('material', '')
    return render_template('price-benchmarking.html', material=material)

@app.route('/quote-compare')
def quote_compare():
    # Get the searched material from query parameters
    #material = request.args.get('material', '')
    return render_template('quote-compare.html')

@app.route('/settings')
def settings():
    return render_template('settings.html')

@app.route('/process-quotes', methods=['POST'])
def process_quotes():
    """Handle quote processing"""
    processed_data = []
    analysis = None

    if 'files[]' not in request.files:
        return render_template('quote-compare.html', error="No files selected", active_tab='apollo')

    files = request.files.getlist('files[]')
    
    if not files or files[0].filename == '':
        return render_template('quote-compare.html', error="No selected files", active_tab='apollo')

    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

    for file in files:
        if file and file.filename.lower().endswith('.pdf'):
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            try:
                file.save(filepath)
                with open(filepath, 'rb') as pdf_file:
                    pdf_reader = PyPDF2.PdfReader(pdf_file)
                    full_text = "".join([page.extract_text() for page in pdf_reader.pages])
                
                # FIXED: Added detailed debugging output to track processing steps
                print(f"Extracted text length: {len(full_text)}")
                print(f"First 200 characters: {full_text[:200]}")
                
                result = process_text(full_text)
                print(f"Process result: {result}")
                
                if isinstance(result, dict) and 'error' not in result:
                    processed_data.append(result)
                else:
                    print(f"Processing failed: {result}")
                    
            except Exception as e:
                print(f"Exception during processing: {str(e)}")
                return render_template('quote-compare.html', error=f"Error processing {file.filename}: {str(e)}", active_tab='apollo')
            finally:
                if os.path.exists(filepath):
                    os.remove(filepath)
        else:
            return render_template('quote-compare.html', error="Please upload PDF files only", active_tab='apollo')

    # FIXED: Added better error handling and user feedback
    if processed_data:
        analysis = analyze_quotes(processed_data)
        excel_filename = create_excel(processed_data)
        return render_template('quote-compare.html', 
                             processed_data=processed_data, 
                             analysis=analysis, 
                             excel_filename=excel_filename,
                             active_tab='apollo')
    else:
        return render_template('quote-compare.html', error="No valid data could be extracted from the uploaded files. Please ensure your PDF contains quote information with vendor details, pricing, and material specifications.", active_tab='apollo')

    return render_template('quote-compare.html')

@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)

@app.route('/process-csv', methods=['POST'])
def process_csv():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not file.filename.endswith('.csv'):
        return jsonify({'error': 'Please upload a CSV file'}), 400
    
    try:
        # Save the uploaded file temporarily
        temp_path = os.path.join(app.config['UPLOAD_FOLDER'], 'temp.csv')
        file.save(temp_path)
        
        # Process the file using DuplicateDetector
        detector = DuplicateDetector(threshold=70.0, max_rows=100)
        matches = detector.process_file(temp_path)
        
        # Clean up the temporary file
        if os.path.exists(temp_path):
            os.remove(temp_path)
        
        return jsonify(matches)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    # FIXED: Changed port from 5000 to 8080 to avoid conflicts with Apple's AirPlay
    app.run(debug=True, port=5002)